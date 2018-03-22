# -*- coding: utf-8 -*-

import os
import json
import requests
import datetime
import time 
from pyquery import PyQuery as pq
import xlwt
from openpyxl import Workbook
from openpyxl import load_workbook
import ssl
import traceback
from bs4 import BeautifulSoup
from lxml import etree
import re
import random
import numpy as np


hds=[{'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.221 Safari/537.36 SE 2.X MetaSr 1.0'},\
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36'},\
    {'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
    {'User-Agent':'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'},\
    {'User-Agent':'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:34.0) Gecko/20100101 Firefox/34.0'},\
    {'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/44.0.2403.89 Chrome/44.0.2403.89 Safari/537.36'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50'},\
    {'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50'},\
    {'User-Agent':'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1'},\
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11'},\
    {'User-Agent':'Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11'},\
    {'User-Agent':'Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11'}]
   
cookies = [{'cookie':'PHPSESSID=tqm22peu5splfnvmn0hpg4lfu5; __guid=90162694.3053780365895112000.1519868913096.387; __q__=1519868911314; _currentUrl_=%2FLoo; __DC_monitor_count=2; __DC_gid=90162694.265178359.1519868913097.1519868914953.3; __DC_sid=90162694.3252704793331246000.1519868913094.9783'},\
          {'cookie':'td_cookie=168205539; PHPSESSID=74ek1pf4qgo3eqd8rk31b8ao90; __DC_sid=90162694.3427863339221094400.1519869151657.8533; __guid=90162694.1790260288613740500.1519869151660.7058; __q__=1519869149544; _currentUrl_=%2FLoo; __DC_monitor_count=2; __DC_gid=90162694.97657791.1519869151659.1519869153252.3'},\
          {'cookie':'__DC_gid=90162694.778235822.1519871167829.1519871171971.3;__DC_monitor_count=2;__DC_sid=90162694.1795709667954440000.1519871167816.0322;__guid=90162694.2409439498319423500.1519871167828.896;__q__	=1519871168187;_currentUrl_=/Loo'}]

def get_info_list(url):
    info=[] 
    count=0
    for pageIndex in range(1,5000):
        count=count+1
        try:
            print('正在爬取第'+str(pageIndex)+'页') 
            json_url =url+'/index/p/'+str(pageIndex)             
            req = requests.get(json_url,headers=hds[random.randint(0,len(hds)-1)],cookies=cookies[random.randint(0,len(cookies)-1)]).text
            soup = BeautifulSoup(req,'lxml')
            selector=etree.HTML(req)       
            contents = selector.xpath('//ul[@class="loopListBottom"]/li')  
            for content in contents:
                try:
                    for i in range(0,len(contents)):
                        if content.xpath('//dl/dd[1]/a'):
                            mfrsname=content.xpath('//dl/dd[1]/a/text()')[i]                          
                            mfrsurl='http://loudong.360.cn'+content.xpath('//dl/dd[1]/a/@href')[i]
                        else:
                            None
                        if content.xpath('//dl/dd[1]/span[3]'):
                            securitytype=content.xpath('//dl/dd[1]/span[3]/text()')[i]
                           
                        else:
                            None
                        if content.xpath('//dl/dd[2]/strong'):
                            importance=content.xpath('//dl/dd[2]/strong/text()')[i]
                           
                        else:
                            None
                        if content.xpath('//dl/dd[2]/span'):
                            type1=content.xpath('//dl/dd[2]/span/text()')[i]
                            
                        else:
                            None                        
                        if content.xpath('//dl/dd[2]/em'):
                            date1=content.xpath('//dl/dd[2]/em/text()')[i]
                            
                        else:
                            None             
                        info.append([mfrsname,mfrsurl,securitytype,importance,type1,date1])
                        print (info)
                    break 
                except:
                    break             
        except:
            print('正在爬取第'+str(pageIndex)+'页出现错误') 
            time.sleep(30)
            continue 
                    
        if count==500:
           content=fill_content(info)
           info=[]
           count=0
         
    return info

    
def fill_content(info):
    s=0
    ws = wb.create_sheet('s')     
    ws.append(['厂商名称','厂商链接','漏洞名称','风险等级','漏洞类型','提交时间'])
    s=s+1
    try:
        for info1 in info:
            ws.append([info1[0],info1[1],info1[2],info1[3],info1[4],info1[5]])
            wb.save(file_name)
    except:
        wb.save(file_name)
    return file_name
   
    
    
    
print(str(datetime.datetime.now()))
file_name = 'E:\code\Butianloudong.xlsx'
wb=Workbook()
ws = wb.active
wb.save(file_name)
url = 'https://butian.360.cn/Loo'
info = get_info_list(url)  
print(str(datetime.datetime.now()))

