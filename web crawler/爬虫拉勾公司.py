import os
import json
import requests
import datetime
import time 
import xlwt
from pyquery import PyQuery as pq
from openpyxl import Workbook
from openpyxl import load_workbook
def get_company_list(areaId):
    company_list=[]
    for pageIndex in range(1,50):
        print('正在爬取第'+str(pageIndex)+'页') 
        time.sleep(30)#暂停30s再爬
        json_url = 'https://www.lagou.com/gongsi/'+areaId+'-0-38.json'
        #模拟浏览器爬虫
        my_headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Host': 'www.lagou.com',
        'Origin': 'https://www.lagou.com',
        'Referer': 'https://www.lagou.com/gongsi/3-0-38?isShowMoreIndustryField=true',
        'Connection':'keep-alive',
        'Accept-Encoding':'gzip, deflate, br',
        'X-Requested-With':'XMLHttpRequest',
        'Content-Type':'application/x-www-form-urlencoded; charset=UTF-8'
        }
        cookies = {'cookie':'JSESSIONID=ABAAABAAAFCAAEG49E565EBAC3AD84B8555E9C9780F3C9B; _gat=1; user_trace_token=20171010174534-c3442683-ad9f-11e7-945a-5254005c3644; PRE_UTM=; PRE_HOST=; PRE_SITE=; PRE_LAND=https%3A%2F%2Fwww.lagou.com%2F; LGUID=20171010174534-c3442984-ad9f-11e7-945a-5254005c3644; index_location_city=%E4%B8%8A%E6%B5%B7; TG-TRACK-CODE=index_company; _gid=GA1.2.539866734.1507628736; _ga=GA1.2.28091358.1507628736; Hm_lvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1507628736; Hm_lpvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1507628755; LGSID=20171010174534-c34427fe-ad9f-11e7-945a-5254005c3644; LGRID=20171010174553-ce9e07ec-ad9f-11e7-945a-5254005c3644'}
        param = {'first': 'false', 'pn': str(pageIndex), 'sortField': '0', 'havemark': '0'} #访问参数
        r = requests.post(json_url,params=param,cookies=cookies,headers=my_headers) #requsets请求
        msg = r.json()
        info = msg['result']
        print (info)
        for company in info:
            company_list.append([company['city'],company['companyFeatures'],company['companyFullName'],company['companyShortName'],company['financeStage'],company['industryField'],company['interviewRemarkNum'],company['positionNum'],company['processRate']])
            print (json.dumps(company_list, ensure_ascii=False, indent=2)) 
            
    return company_list
   

def write_file(file_name):
    list = []
    wb = Workbook()
    ws = wb.active
    url = 'https://www.lagou.com/gongsi/3-0-38?isShowMoreIndustryField=true'
    wb.create_sheet(title = '上海')
    wb.save(file_name)
    areaId = '3'
    company_list = get_company_list(areaId)
    ws = wb.get_sheet_by_name(company_list[0][0])
    ws.append(['城市名称','公司简介','公司全称','公司简称','财务状况','行业','评价打分','在招职位数量','简历处理率'])
    for company in company_list:
        ws.append([company[0],company[1],company[2],company[3],company[4],company[5],company[6],company[7],company[8]])
        wb.save(file_name)
                   
file_name = 'shanghai1.xlsx'
print(str(datetime.datetime.now()))
write_file(file_name)
print(str(datetime.datetime.now()))