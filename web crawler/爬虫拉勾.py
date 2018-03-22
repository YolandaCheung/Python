import os
import json
import requests
import datetime
import time 
import xlwt
from pyquery import PyQuery as pq
from openpyxl import Workbook
from openpyxl import load_workbook
import ssl
import traceback

#获取城市ID列表
def get_cityId_list(url):
    city_list = []
    html = pq(url= url)
    for areaId in html.find('#filterCollapse').find('div[class="has-more workcity"]').eq(0).find('div[class="more more-positions"]').find("a[data-lg-tj-cid='idnull']"):
        aId = pq(areaId).attr("data-id")
        if(aId=='1'):
            continue
        city_list.append(aId)
    return city_list

    
#获取城市名称
def get_city_name_list(u):
    city_name_list = []
    html = pq(url=url)
    for areaId in html.find('#filterCollapse').find('div[class="has-more workcity"]').eq(0).find('div[class="more more-positions"]').find("a[data-lg-tj-cid='idnull']"):
        area_name=pq(areaId).html()
        if area_name=="u'\u5168\u56fd'":
            continue
        city_name_list.append(area_name)
    return city_name_list

    
#获取公司信息
def get_company_list(areaId):
    company_list=[]
    for pageIndex in range(1,30):
        print('正在爬取第'+str(pageIndex)+'页') 
        time.sleep(30)#暂停30s再爬
        json_url = 'https://www.lagou.com/gongsi/'+areaId+'-0-38.json'
        #模拟浏览器爬虫
        my_headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Host': 'www.lagou.com',
        'Origin': 'https://www.lagou.com',
        'Connection':'keep-alive',
        'Accept-Encoding':'gzip, deflate, br',
        'X-Requested-With':'XMLHttpRequest',
        'Content-Type':'application/x-www-form-urlencoded; charset=UTF-8'
        }
        cookies = {'cookie':'user_trace_token=20171003120856-931a4b92-a7f0-11e7-b901-525400f775ce; LGUID=20171003120856-931a4e62-a7f0-11e7-b901-525400f775ce; index_location_city=%E5%85%A8%E5%9B%BD; TG-TRACK-CODE=index_company; _gat=1; PRE_UTM=; PRE_HOST=; PRE_SITE=https%3A%2F%2Fwww.lagou.com%2Fgongsi%2F; PRE_LAND=https%3A%2F%2Fwww.lagou.com%2Fgongsi%2F0-0-38%3FisShowMoreIndustryField%3Dtrue; JSESSIONID=ABAAABAAAIAACBIE4AF9C33510F52118E9144758815B4C8; LGSID=20171003202247-90e4f92c-a835-11e7-b928-525400f775ce; LGRID=20171003202312-9fb4c032-a835-11e7-b928-525400f775ce; Hm_lvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1507003738,1507022122; Hm_lpvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1507033395; _ga=GA1.2.1999978111.1507003739; _gid=GA1.2.1623212744.1507003739'}
        param = {'first': 'false', 'pn': str(pageIndex), 'sortField': '0', 'havemark': '0'} #访问参数
        r = requests.post(json_url,params=param,cookies=cookies,headers=my_headers) #requsets请求  
        msg = r.json()
        print (msg)
        info = msg['result']
        print (info)
        for company in info:
            company_list.append([company['city'],company['companyFeatures'],company['companyFullName'],company['companyShortName'],company['finaceStage'],company['industryField']])
            
            
    return company_list
   
def write_file(fileName):
    list = []
    wb = Workbook()
    ws = wb.active
    url= 'https://www.lagou.com/gongsi/'
    area_name_list = get_city_name_list(url)
    print (area_name_list)
    for area_name in area_name_list:
        wb.create_sheet(title = area_name)
        file_name = fileName+'.xlsx'
        wb.save(file_name)
    areaId_list = get_cityId_list(url)
    print (areaId_list)
    for areaId in areaId_list:
        company_list = get_company_list(areaId)
        print('正在爬取----->****'+company_list[0][0]+'****公司列表')
        wb1 = load_workbook(file_name)
        ws = wb1.get_sheet_by_name(company_list[0][0])
        ws.append(['城市名称','公司简介','公司全称','公司简称','财务状况','行业'])
        for company in company_list:
            ws.append([company[0],company[1],company[2],company[3],company[4],company[5]])
            wb1.save(file_name)
                   
file_name = 'palagou'
print(str(datetime.datetime.now()))
write_file(file_name)
print(str(datetime.datetime.now()))