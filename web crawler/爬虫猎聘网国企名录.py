# -*- coding:utf-8 -*-
import requests
from bs4 import BeautifulSoup
from lxml import etree
import re
import random
import time
import json
from openpyxl import Workbook
from openpyxl import load_workbook
import xlwt
import datetime 

hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
    {'User-Agent':'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'},\
    {'User-Agent':'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:34.0) Gecko/20100101 Firefox/34.0'},\
    {'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/44.0.2403.89 Chrome/44.0.2403.89 Safari/537.36'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50'},\
    {'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50'},\
    {'User-Agent':'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1'},\
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11'},\
    {'User-Agent':'Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11'},\
    {'User-Agent':'Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11'}]

def city_list(base_url):
    href_list = []
    base_url = 'https://www.liepin.com/company/000-000-040/'
    req = requests.get(base_url).text
    soup = BeautifulSoup(req,'lxml')
    hrefs= soup.select('#region > div.wrap > div.top-bar > div.company-place > div.place-name > a')
    for href in hrefs:
        href_list.append(href['href'])
    href_list.pop()
    return href_list

def url_page(base_url):
    url_list=[]
    req = requests.get(base_url,headers=hds[random.randint(0,len(hds)-1)]).text
    try:
        pages = re.findall('<span.*?"addition">(.*?)<span.*?"redirect">',req,re.S)[0].replace('共','').replace('页','')
        if pages:
            for i in range(0,int(pages)):
                urls = str(base_url) + 'pn' + str(i)
                url_list.append(urls)
        else:
            urls=url
    except:
        urls= base_url
        url_list.append(urls)               
    return url_list

def get_company_url(base_url):
    b=[]
    for urls in url_list:
        req = requests.get(urls,headers=hds[random.randint(0,len(hds)-1)]).text
        soup = BeautifulSoup(req,'lxml')
        companies = soup.select('#region > div.wrap > div.company-list.clearfix > div > div.item-top.clearfix > div > p.company-name > a')
        for company in companies:
            company_url= company['href']
            b.append(company_url)                 
    return b 

def get_company_info(base_url):
    company_info=[]   
    companyname =[]
    industry=[]
    company_size=[]
    company_date=[]
    company_capital=[]
    company_field=[]
    company_position=[]
    for company_url in b:
        try: 
            req = requests.get(company_url,headers=hds[random.randint(0,len(hds)-1)]).text
            soup = BeautifulSoup(req,'lxml')
            selector = etree.HTML(req)            
            
            if selector.xpath('//*[@id="company"]/div[2]/section/div/h1/text()'):
                companyname = selector.xpath('//*[@id="company"]/div[2]/section/div/h1/text()')[0] 
            else:
                None
             
            if selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[1]/li[1]/a/text()'):
                industry= selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[1]/li[1]/a/text()')[0]
            else:
                None
            
            if selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[1]/li[3]/text()'):
                company_size= selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[1]/li[2]/text()')[0]
            else:
                None
            
            if selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[2]/li[2]/text()'):
                company_date= selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[2]/li[2]/text()')[0]
            else:
                None 
              
            if selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[2]/li[3]/text()'):
                company_capital= selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[2]/li[3]/text()')[0] 
            else:
                None
          
            if selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[2]/li[5]/text()'):
                company_field = selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[2]/li[5]/text()')[0]             
            else:
                None 
            if selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[1]/li[3]/@data-point')[0]:
                company_position = selector.xpath('//*[@id="company"]/div[2]/div/aside/div[2]/ul[1]/li[3]/@data-city')[0]
            else:
                None 
            company_info.append([companyname,industry,company_size,company_date,company_capital,company_field,company_position])
        except:
            continue 
    return company_info

wb = Workbook()
base_url='https://www.liepin.com/company/090-000-040/'
url_list=url_page(base_url)
print (url_list)
b= get_company_url(base_url)
print (b)
company_info=get_company_info(base_url)
file_name = 'lpfj1.xlsx'
wb.create_sheet(title = '公司信息')
wb.save(file_name)
ws = wb.get_sheet_by_name(company_info[0][0])
ws.append(['公司名称','行业','公司规模','注册时间','注册金额','经营范围','地址'])
for company in company_info:
    try: 
        if company[6]:
            ws.append([company[0],company[1],company[2],company[3],company[4],company[5],company[6]])
        else:
            ws.append([company[0],company[1],company[2],company[3],company[4],company[5]])
        wb.save(file_name)